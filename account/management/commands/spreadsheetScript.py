# import load_workbook
import json
from email_validator import validate_email, EmailNotValidError
from openpyxl import load_workbook
import datetime

flaggedRows=[]

def main():
    path=input("Spreadsheet Directory: ")
    dataType=input("Input data type")
    checkSheet(path,dataType)

def checkSheet(path,dataType):
    sheet = load_workbook(path).active
    for currentRow in range(2, sheet.max_row + 1):
        for currentColumn in range(1, sheet.max_column + 1):
            if(dataType=="Student"):
                print(sheet.cell(row=currentRow, column=currentColumn).value)
                currentValue=sheet.cell(row=currentRow, column=currentColumn).value
                if(currentColumn in [1,2,3,5,6]):
                    if(nameChecker(currentValue)==False):
                        flaggedRows.append([currentRow, "name error"])
                if(currentColumn == 4):
                   #dateTimeValidator(currentValue, '%m/%d/%Y')
                    print("date validation")
                if(currentColumn == 7):
                    emailValidator(currentValue)
                if(currentColumn == 8):
                    if(isinstance(currentValue,int)==False):
                        flaggedRows.append([currentRow,"Grade error"])


    print(flaggedRows)

def dateTimeValidator(dateTime, format):
    try:
        dateTime.datetime.strptime(dateTime, format)
    except ValueError:
        return False

def nameChecker(name):
    if name.replace(" ","").isalpha():
        return True
    else:
        return False

def emailValidator(email):
    try:
        validate_email(email).email
    except EmailNotValidError as e:
        # email is not valid, exception message is human-readable
        return str(e)

if __name__ == '__main__':
    main()